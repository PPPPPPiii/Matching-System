from __future__ import annotations

import csv
import hashlib
import re
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

from openpyxl import load_workbook

from .models import Participant
from .models import parse_bool


class GoogleFormImportError(ValueError):
    pass


def _normalize_header(value: str) -> str:
    raw_tokens = re.findall(r"[a-z0-9]+", value.lower())
    token_aliases = {
        "countries": "country",
        "nationalities": "nationality",
        "cultures": "culture",
        "citizens": "citizen",
        "women": "woman",
        "men": "man",
        "scholars": "scholar",
        "students": "student",
    }
    tokens = [token_aliases.get(token, token) for token in raw_tokens]
    return " ".join(tokens)


def _record_key(source: str, row: Dict[str, str]) -> str:
    normalized = "|".join(
        f"{key}:{(value or '').strip().lower()}" for key, value in sorted(row.items())
    )
    raw = f"{source}|{normalized}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _contains_all(tokens: set[str], words: Sequence[str]) -> bool:
    return all(word in tokens for word in words)


def _detect_column_indices(headers: Sequence[str]) -> Dict[str, int]:
    normalized = [_normalize_header(header) for header in headers]
    token_sets = [set(value.split()) for value in normalized]

    # "Name" can be plain full-name or question-style phrasing that includes
    # first+last in one column (e.g., "What is your first and last name?").
    full_name_idx = -1
    full_name_score = -1
    for index, tokens in enumerate(token_sets):
        if "name" not in tokens:
            continue
        has_split_name_markers = bool(
            tokens.intersection({"first", "last", "given", "family", "surname"})
        )
        if has_split_name_markers and not (
            {"first", "last", "name"}.issubset(tokens)
        ):
            continue
        if {"full", "name"}.issubset(tokens):
            score = 3
        elif {"first", "last", "name"}.issubset(tokens):
            score = 2
        else:
            score = 1
        if score > full_name_score:
            full_name_score = score
            full_name_idx = index

    def find_column(candidates: Sequence[Sequence[str]]) -> int:
        best_index = -1
        best_score = -1
        for index, tokens in enumerate(token_sets):
            for candidate in candidates:
                if _contains_all(tokens, candidate):
                    score = len(candidate)
                    if score > best_score:
                        best_score = score
                        best_index = index
        return best_index

    first_name_idx = find_column((("first", "name"), ("given", "name")))
    last_name_idx = find_column((("last", "name"), ("family", "name"), ("surname",)))
    country_idx = find_column(
        (
            ("country", "citizen"),
            ("country", "citizenship"),
            ("country",),
            ("citizenship",),
        )
    )
    culture_idx = find_column(
        (
            ("nationality",),
            ("nationalities",),
            ("culture",),
            ("identified", "as"),
            ("identify", "as"),
        )
    )
    gender_idx = find_column(
        (
            ("gender",),
            ("sex",),
            ("woman",),
            ("female",),
            ("man",),
            ("male",),
        )
    )
    emory_idx = find_column(
        (
            ("emory", "student"),
            ("emory",),
            ("student", "scholar"),
            ("student",),
            ("scholar",),
        )
    )
    first_time_idx = find_column(
        (
            ("first", "time"),
            ("first", "timer"),
            ("attendance", "experience"),
            ("attended", "before"),
            ("returning",),
        )
    )
    age_idx = find_column((("age",),))

    required_missing: list[str] = []
    if full_name_idx < 0 and (first_name_idx < 0 or last_name_idx < 0):
        required_missing.append("name (full name or first+last name)")
    if country_idx < 0:
        required_missing.append("country/citizenship")
    if culture_idx < 0:
        required_missing.append("nationality/culture identified as")
    if gender_idx < 0:
        required_missing.append("gender")
    if emory_idx < 0:
        required_missing.append("Emory student")
    if first_time_idx < 0:
        required_missing.append("first-time or attended-before")
    if required_missing:
        raise GoogleFormImportError(
            "Unable to detect required columns from first header row: "
            + ", ".join(required_missing)
        )

    return {
        "full_name": full_name_idx,
        "first_name": first_name_idx,
        "last_name": last_name_idx,
        "country": country_idx,
        "culture": culture_idx,
        "gender": gender_idx,
        "is_emory_student": emory_idx,
        "first_time": first_time_idx,
        "age": age_idx,
    }


def _read_csv_matrix(path: Path) -> List[List[str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        reader = csv.reader(fh)
        return [[(cell or "").strip() for cell in row] for row in reader]


def _read_xlsx_matrix(path: Path) -> List[List[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    matrix: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        matrix.append([("" if cell is None else str(cell)).strip() for cell in row])
    wb.close()
    return matrix


def _read_sheet_matrix(file_path: str) -> List[List[str]]:
    path = Path(file_path)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv_matrix(path)
    if suffix == ".xlsx":
        return _read_xlsx_matrix(path)
    raise GoogleFormImportError("Unsupported file type. Please upload a .csv or .xlsx file.")


def _row_to_dict(headers: Sequence[str], row: Sequence[str]) -> Dict[str, str]:
    values = list(row) + [""] * max(len(headers) - len(row), 0)
    return {headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))}


def _cell(row: Sequence[str], idx: int) -> str:
    if idx < 0:
        return ""
    if idx >= len(row):
        return ""
    return (row[idx] or "").strip()


def _parse_first_time_value(header_text: str, raw_value: str) -> bool:
    normalized_value = _normalize_header(raw_value)
    value_tokens = set(normalized_value.split())
    if value_tokens.intersection({"yes", "y", "true", "1"}):
        value = True
    elif value_tokens.intersection({"no", "n", "false", "0"}):
        value = False
    elif "first" in value_tokens and "time" in value_tokens:
        value = True
    elif "not" in value_tokens and "first" in value_tokens:
        value = False
    else:
        value = parse_bool(raw_value)

    tokens = set(_normalize_header(header_text).split())
    attended_before_style = "before" in tokens or "experience" in tokens or "returning" in tokens
    return (not value) if attended_before_style else value


def _parse_student_or_scholar_value(raw_value: str) -> bool:
    normalized_value = _normalize_header(raw_value)
    value_tokens = set(normalized_value.split())

    if value_tokens.intersection({"yes", "y", "true", "1"}):
        return True
    if value_tokens.intersection({"no", "n", "false", "0"}):
        return False

    # Phrase-style survey responses such as:
    # "No, I am not a university student or scholar"
    # "Yes, I am an Emory undergrad student"
    has_positive_student_marker = bool(
        value_tokens.intersection({"student", "scholar", "undergrad", "undergraduate", "grad", "graduate"})
    )
    has_negative_marker = "not" in value_tokens or "no" in value_tokens
    if has_positive_student_marker:
        return not has_negative_marker

    return parse_bool(raw_value)


def parse_uploaded_sheet(
    file_path: str,
    *,
    default_age: int = 21,
    skip_incomplete_rows: bool = True,
) -> List[Tuple[str, Participant]]:
    matrix = _read_sheet_matrix(file_path)
    if not matrix:
        return []

    headers = [(cell or "").strip() for cell in matrix[0]]
    if not any(headers):
        raise GoogleFormImportError("First row must contain header names.")

    columns = _detect_column_indices(headers)
    source = str(Path(file_path).resolve())
    parsed: List[Tuple[str, Participant]] = []
    skipped_incomplete = 0

    for row_index, row in enumerate(matrix[1:], start=2):
        if not any((cell or "").strip() for cell in row):
            continue

        row_dict = _row_to_dict(headers, row)
        if columns["full_name"] >= 0:
            name = _cell(row, columns["full_name"])
        else:
            first_name = _cell(row, columns["first_name"])
            last_name = _cell(row, columns["last_name"])
            name = f"{first_name} {last_name}".strip()
        if not name:
            if skip_incomplete_rows:
                skipped_incomplete += 1
                continue
            raise GoogleFormImportError(f"Row {row_index}: missing participant name.")

        raw_first_time = _cell(row, columns["first_time"])
        if not raw_first_time:
            if skip_incomplete_rows:
                skipped_incomplete += 1
                continue
            raise GoogleFormImportError(
                f"Row {row_index}: missing first-time/attended-before value."
            )
        raw_country = _cell(row, columns["country"])
        raw_culture = _cell(row, columns["culture"])
        raw_gender = _cell(row, columns["gender"])
        raw_emory_status = _cell(row, columns["is_emory_student"])
        if not (raw_country and raw_culture and raw_gender and raw_emory_status):
            if skip_incomplete_rows:
                skipped_incomplete += 1
                continue
            raise GoogleFormImportError(
                f"Row {row_index}: missing one or more required fields "
                "(country/culture/gender/emory-student)."
            )
        header_for_first_time = headers[columns["first_time"]]
        first_time = _parse_first_time_value(header_for_first_time, raw_first_time)

        age_text = _cell(row, columns["age"]) if columns["age"] >= 0 else ""
        age = default_age if not age_text else int(age_text)

        participant = Participant.from_signup(
            name=name,
            age=age,
            is_emory_student=_parse_student_or_scholar_value(raw_emory_status),
            gender=raw_gender,
            attendance_experience=(not first_time),
            ethnicity=raw_country,
            culture=raw_culture,
        )
        parsed.append((_record_key(source, row_dict), participant))

    parse_uploaded_sheet.last_skipped_incomplete = skipped_incomplete
    return parsed


# Backward compatible aliases
parse_google_form_rows = parse_uploaded_sheet
parse_uploaded_sheet.last_skipped_incomplete = 0
